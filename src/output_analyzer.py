"""
Output analyzer module for generating enhanced analysis reports.

This module provides:
- Enhanced distance analysis with optimized distances
- Enhanced coordinate analysis with optimized coordinates and error metrics
- Sorted and organized output tables
- Comprehensive error analysis per point
"""

import pandas as pd
import numpy as np
from typing import Dict, Tuple, List
from .distance_calculator import DistanceCalculator


class OutputAnalyzer:
    """
    Analyzes and generates enhanced output reports for coordinate estimation results.
    
    This class provides methods to:
    - Generate enhanced distance analysis with optimized distances
    - Generate enhanced coordinate analysis with error metrics
    - Create sorted and organized output tables
    - Calculate per-point error statistics
    """
    
    def __init__(self):
        """Initialize the output analyzer."""
        self.distance_calculator = DistanceCalculator()
    
    def create_enhanced_distance_analysis(self, 
                                        original_distances: pd.DataFrame,
                                        optimized_coordinates: Dict[str, Tuple[float, float]]) -> pd.DataFrame:
        """
        Create enhanced distance analysis with optimized distances.
        
        Args:
            original_distances: Original distance measurements DataFrame
            optimized_coordinates: Dictionary of optimized coordinates
            
        Returns:
            Enhanced DataFrame with sorted columns and optimized distances
        """
        # Create a copy of the original data
        enhanced_df = original_distances.copy()
        
        # Add optimized distance column
        optimized_distances = []
        for _, row in enhanced_df.iterrows():
            point1 = row['First point ID']
            point2 = row['Second point ID']
            
            # Get optimized coordinates
            coords1 = optimized_coordinates[point1]
            coords2 = optimized_coordinates[point2]
            
            # Calculate optimized distance
            optimized_distance = self.distance_calculator.euclidean_distance(coords1, coords2)
            optimized_distances.append(optimized_distance)
        
        enhanced_df['Optimized Distance'] = optimized_distances
        
        # Add distance error column (absolute difference between measured and optimized)
        distance_errors = []
        for _, row in enhanced_df.iterrows():
            measured_distance = row['Distance']
            optimized_distance = row['Optimized Distance']
            error = abs(measured_distance - optimized_distance)
            distance_errors.append(error)
        
        enhanced_df['Distance Error'] = distance_errors
        
        # Sort the first two columns alphabetically in each row
        for idx, row in enhanced_df.iterrows():
            point1 = row['First point ID']
            point2 = row['Second point ID']
            
            # Sort alphabetically
            if point1 > point2:
                enhanced_df.at[idx, 'First point ID'] = point2
                enhanced_df.at[idx, 'Second point ID'] = point1
        
        # Sort rows alphabetically (first column primary, second column secondary)
        enhanced_df = enhanced_df.sort_values(['First point ID', 'Second point ID'])
        
        return enhanced_df
    
    def create_enhanced_coordinate_analysis(self,
                                          original_coordinates: pd.DataFrame,
                                          optimized_coordinates: Dict[str, Tuple[float, float]],
                                          distances: Dict[Tuple[str, str], float]) -> pd.DataFrame:
        """
        Create enhanced coordinate analysis with optimized coordinates and error metrics.
        
        Args:
            original_coordinates: Original coordinates DataFrame
            optimized_coordinates: Dictionary of optimized coordinates
            distances: Dictionary of distance measurements
            
        Returns:
            Enhanced DataFrame with optimized coordinates and error metrics
        """
        # Create a copy of the original data
        enhanced_df = original_coordinates.copy()
        
        # Add optimized coordinates
        optimized_x = []
        optimized_y = []
        distance_counts = []
        rms_errors = []
        movement_distances = []
        
        for _, row in enhanced_df.iterrows():
            point_id = row['Point ID']
            
            # Get initial coordinates
            initial_x = row['x_guess']
            initial_y = row['y_guess']
            
            # Get optimized coordinates
            opt_x, opt_y = optimized_coordinates[point_id]
            optimized_x.append(opt_x)
            optimized_y.append(opt_y)
            
            # Calculate distance moved from initial to optimized position
            movement_distance = self.distance_calculator.euclidean_distance(
                (initial_x, initial_y), (opt_x, opt_y)
            )
            movement_distances.append(movement_distance)
            
            # Calculate number of measured distances involving this point
            count = self._count_distances_for_point(point_id, distances)
            distance_counts.append(count)
            
            # Calculate RMS error for all measurements involving this point
            rms_error = self._calculate_rms_error_for_point(point_id, optimized_coordinates, distances)
            rms_errors.append(rms_error)
        
        # Add new columns
        enhanced_df['Optimized X'] = optimized_x
        enhanced_df['Optimized Y'] = optimized_y
        enhanced_df['Movement Distance'] = movement_distances
        enhanced_df['Distance Count'] = distance_counts
        enhanced_df['RMS Error'] = rms_errors
        
        # Sort by first column (Point ID) alphabetically
        enhanced_df = enhanced_df.sort_values('Point ID')
        
        return enhanced_df
    
    def _count_distances_for_point(self, point_id: str, 
                                  distances: Dict[Tuple[str, str], float]) -> int:
        """
        Count the number of distance measurements involving a specific point.
        
        Args:
            point_id: The point ID to count distances for
            distances: Dictionary of distance measurements
            
        Returns:
            Number of distance measurements involving this point
        """
        count = 0
        for pair in distances.keys():
            if point_id in pair:
                count += 1
        return count
    
    def _calculate_rms_error_for_point(self, point_id: str,
                                     optimized_coordinates: Dict[str, Tuple[float, float]],
                                     distances: Dict[Tuple[str, str], float]) -> float:
        """
        Calculate RMS error for all distance measurements involving a specific point.
        
        Args:
            point_id: The point ID to calculate RMS error for
            optimized_coordinates: Dictionary of optimized coordinates
            distances: Dictionary of distance measurements
            
        Returns:
            RMS error for all measurements involving this point
        """
        errors = []
        
        for pair, measured_distance in distances.items():
            if point_id in pair:
                # Get the other point in the pair
                other_point = pair[0] if pair[1] == point_id else pair[1]
                
                # Get coordinates for both points
                coords1 = optimized_coordinates[point_id]
                coords2 = optimized_coordinates[other_point]
                
                # Calculate optimized distance
                calculated_distance = self.distance_calculator.euclidean_distance(coords1, coords2)
                
                # Calculate error
                error = measured_distance - calculated_distance
                errors.append(error)
        
        if not errors:
            return 0.0
        
        # Calculate RMS error
        rms_error = np.sqrt(np.mean(np.array(errors) ** 2))
        return float(rms_error)
    
    def generate_enhanced_output(self,
                                original_distances: pd.DataFrame,
                                original_coordinates: pd.DataFrame,
                                optimized_coordinates: Dict[str, Tuple[float, float]],
                                distances: Dict[Tuple[str, str], float]) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """
        Generate complete enhanced output analysis.
        
        Args:
            original_distances: Original distance measurements DataFrame
            original_coordinates: Original coordinates DataFrame
            optimized_coordinates: Dictionary of optimized coordinates
            distances: Dictionary of distance measurements
            
        Returns:
            Tuple of (enhanced_distances_df, enhanced_coordinates_df)
        """
        # Create enhanced distance analysis
        enhanced_distances = self.create_enhanced_distance_analysis(
            original_distances, optimized_coordinates
        )
        
        # Create enhanced coordinate analysis
        enhanced_coordinates = self.create_enhanced_coordinate_analysis(
            original_coordinates, optimized_coordinates, distances
        )
        
        return enhanced_distances, enhanced_coordinates
    
    def save_enhanced_output(self,
                           enhanced_distances: pd.DataFrame,
                           enhanced_coordinates: pd.DataFrame,
                           output_path: str):
        """
        Save enhanced output to Excel file with formatting and sorting.
        
        Args:
            enhanced_distances: Enhanced distance analysis DataFrame
            enhanced_coordinates: Enhanced coordinate analysis DataFrame
            output_path: Path to save the Excel file
        """
        # Sort the DataFrames before saving
        enhanced_distances_sorted = enhanced_distances.sort_values('Distance Error', ascending=False)
        enhanced_coordinates_sorted = enhanced_coordinates.sort_values('Movement Distance', ascending=False)
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Save the sorted DataFrames
            enhanced_distances_sorted.to_excel(writer, sheet_name='Enhanced Distances', index=False)
            enhanced_coordinates_sorted.to_excel(writer, sheet_name='Enhanced Coordinates', index=False)
            
            # Get the workbook and worksheets for formatting
            workbook = writer.book
            
            # Format the Enhanced Distances sheet
            if 'Enhanced Distances' in workbook.sheetnames:
                ws_distances = workbook['Enhanced Distances']
                self._format_excel_sheet(ws_distances, "Enhanced Distances")
            
            # Format the Enhanced Coordinates sheet
            if 'Enhanced Coordinates' in workbook.sheetnames:
                ws_coordinates = workbook['Enhanced Coordinates']
                self._format_excel_sheet(ws_coordinates, "Enhanced Coordinates")
        
        print(f"Enhanced output saved to: {output_path}")
        print(f"  - Enhanced Distances sheet sorted by Distance Error (decreasing)")
        print(f"  - Enhanced Coordinates sheet sorted by Movement Distance (decreasing)")
        print(f"  - First row formatted with word wrapping")
    
    def _format_excel_sheet(self, worksheet, sheet_name):
        """
        Format an Excel worksheet with word wrapping for the first row and numeric formatting.
        
        Args:
            worksheet: The openpyxl worksheet object
            sheet_name: Name of the sheet for reference
        """
        from openpyxl.styles import Alignment
        
        # Format the first row (headers) with word wrapping
        for cell in worksheet[1]:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Set all column widths to 10 units
        for column in worksheet.columns:
            column_letter = column[0].column_letter
            worksheet.column_dimensions[column_letter].width = 10
        
        # Format numeric data to 3 decimal places
        for row in worksheet.iter_rows(min_row=2):  # Skip header row
            for cell in row:
                if cell.value is not None:
                    # Check if the cell contains a numeric value
                    try:
                        float_val = float(cell.value)
                        # Format to 3 decimal places
                        cell.number_format = '0.000'
                    except (ValueError, TypeError):
                        # Not a numeric value, leave as is
                        pass
    
    def print_enhanced_summary(self,
                              enhanced_distances: pd.DataFrame,
                              enhanced_coordinates: pd.DataFrame,
                              output_dir: str = None):
        """
        Print a summary of the enhanced analysis and optionally save to file.
        
        Args:
            enhanced_distances: Enhanced distance analysis DataFrame
            enhanced_coordinates: Enhanced coordinate analysis DataFrame
            output_dir: Optional output directory to save summary text file
        """
        # Build the summary text
        summary_lines = []
        summary_lines.append("=== ENHANCED OUTPUT ANALYSIS ===")
        
        summary_lines.append("\nDistance Analysis Summary:")
        summary_lines.append(f"  Total distance measurements: {len(enhanced_distances)}")
        summary_lines.append(f"  Average optimized distance: {enhanced_distances['Optimized Distance'].mean():.3f}")
        summary_lines.append(f"  Distance range: {enhanced_distances['Optimized Distance'].min():.3f} - {enhanced_distances['Optimized Distance'].max():.3f}")
        summary_lines.append(f"  Average distance error: {enhanced_distances['Distance Error'].mean():.3f}")
        summary_lines.append(f"  Maximum distance error: {enhanced_distances['Distance Error'].max():.3f}")
        
        summary_lines.append("\nCoordinate Analysis Summary:")
        summary_lines.append(f"  Total points: {len(enhanced_coordinates)}")
        summary_lines.append(f"  Average distance count per point: {enhanced_coordinates['Distance Count'].mean():.1f}")
        summary_lines.append(f"  Average RMS error per point: {enhanced_coordinates['RMS Error'].mean():.6f}")
        summary_lines.append(f"  Maximum RMS error: {enhanced_coordinates['RMS Error'].max():.6f}")
        summary_lines.append(f"  Average movement distance: {enhanced_coordinates['Movement Distance'].mean():.3f}")
        summary_lines.append(f"  Maximum movement distance: {enhanced_coordinates['Movement Distance'].max():.3f}")
        
        # Show points with highest RMS errors
        high_error_points = enhanced_coordinates.nlargest(3, 'RMS Error')
        summary_lines.append("\nPoints with highest RMS errors:")
        for _, row in high_error_points.iterrows():
            summary_lines.append(f"  {row['Point ID']}: RMS Error = {row['RMS Error']:.6f}")
        
        # Show points with most distance measurements
        high_count_points = enhanced_coordinates.nlargest(3, 'Distance Count')
        summary_lines.append("\nPoints with most distance measurements:")
        for _, row in high_count_points.iterrows():
            summary_lines.append(f"  {row['Point ID']}: {row['Distance Count']} measurements")
        
        # Show points that moved the most
        high_movement_points = enhanced_coordinates.nlargest(5, 'Movement Distance')
        summary_lines.append("\nPoints that moved the most:")
        for _, row in high_movement_points.iterrows():
            point_id = row['Point ID']
            movement = row['Movement Distance']
            initial_x = row['x_guess']
            initial_y = row['y_guess']
            opt_x = row['Optimized X']
            opt_y = row['Optimized Y']
            summary_lines.append(f"  {point_id}: ({initial_x:.1f}, {initial_y:.1f}) -> ({opt_x:.1f}, {opt_y:.1f}) (moved {movement:.3f})")
        
        # Show distance measurements with highest errors
        high_error_distances = enhanced_distances.nlargest(5, 'Distance Error')
        summary_lines.append("\nDistance measurements with highest errors:")
        for _, row in high_error_distances.iterrows():
            point1 = row['First point ID']
            point2 = row['Second point ID']
            measured = row['Distance']
            optimized = row['Optimized Distance']
            error = row['Distance Error']
            summary_lines.append(f"  {point1}-{point2}: {measured:.3f} -> {optimized:.3f} (error: {error:.3f})")
        
        # Print to console
        print("\n" + "\n".join(summary_lines))
        
        # Save to file if output directory is provided
        if output_dir:
            import os
            summary_file_path = os.path.join(output_dir, "enhanced_analysis_summary.txt")
            try:
                with open(summary_file_path, 'w') as f:
                    f.write("\n".join(summary_lines))
                print(f"\nEnhanced analysis summary saved to: {summary_file_path}")
            except Exception as e:
                print(f"Warning: Could not save summary to file: {str(e)}") 